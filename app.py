from flask import Flask, render_template, request, redirect, url_for,jsonify
import openpyxl
import pandas as pd
import boto3
from datetime import datetime,timedelta
import os
import logging

app = Flask(__name__)
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("flask_app.log"),
        logging.StreamHandler()
    ]
)

@app.route('/error')
def error():
    logging.error("An error occurred!")
    return "Error route!", 500


# @app.route('/', methods=['GET', 'POST'])
# def login_form():
#     return render_template('index.html')

@app.route('/user_dashboard', methods=['GET', 'POST'])
def transaction_form():
    success = False
    stock_insufficient=False
    if request.method == 'POST':
  
        brand = request.form.get('brand')
        transaction_type = request.form.get('transaction_type')
        product_name=request.form.get('product_name')
        units = request.form.get('units')
        purchaser = request.form.get('purchaser')
        remarks = request.form.get('remarks')
        date = request.form.get('date')
      
        

        if not transaction_type:   
            transaction_type = request.form.get('transaction_type_choice')
            return render_template(
                'form.html', brand=brand.lower(), product_name=product_name,units=units, date=date, transaction_type=transaction_type, remarks=remarks
            )

        
        stock_insufficient=add_data(brand,product_name,transaction_type,units,purchaser,date,remarks)
        success=True

        if stock_insufficient:
            success=False  

    return render_template('form.html', success=success,stock_insufficient=stock_insufficient)

def add_data(brand,product_name,transaction_type,units,purchaser,date,remarks):
    curr_month = datetime.now().strftime("%B_%Y")
    workbook=openpyxl.load_workbook(f'artifacts/data/stock_data/{curr_month}.xlsx')
    sheet=workbook.active

    header=1
    headers={cell.value:cell.column for cell in sheet[header]}

    new_workbook=openpyxl.load_workbook(f'artifacts/data/transaction_log/{curr_month}.xlsx')
    new_sheet=new_workbook[brand]

    if transaction_type=='sold':
        for rows in sheet.iter_rows(min_row=2,max_row=sheet.max_row):
        
            if rows[7].value==brand and rows[1].value==product_name:
                code=rows[headers['Code']-1].value
                closing_stock=rows[headers['Closing Stock']-1].value
                if int(units)>int(closing_stock):
                    return True
                sheet.cell(row=rows[0].row,column=headers['Sale Out']).value=int(rows[headers['Sale Out']-1].value)+int(units)
                sheet.cell(row=rows[0].row,column=headers['Closing Stock']).value=int(rows[headers['Closing Stock']-1].value)-int(units)
                new_sheet.append({1:code,2:product_name,3:int(units),5:purchaser,6:date,7:remarks})

    else:
        for rows in sheet.iter_rows(min_row=2,max_row=sheet.max_row):
        
            if rows[7].value==brand and rows[1].value==product_name:
                code=rows[headers['Code']-1].value
        
                sheet.cell(row=rows[0].row,column=headers['Purchase']).value=int(rows[headers['Purchase']-1].value)+int(units)
                sheet.cell(row=rows[0].row,column=headers['Closing Stock']).value=int(rows[headers['Closing Stock']-1].value)+int(units)
                new_sheet.append({1:code,2:product_name,4:int(units),6:date,7:remarks})


    

    workbook.save(f'artifacts/data/stock_data/{curr_month}.xlsx')
    new_workbook.save(f'artifacts/data/transaction_log/{curr_month}.xlsx')

    return False

@app.route('/')
def index():
    return render_template('admin_dashboard.html', products=None, brand_name=None)

@app.route('/search', methods=['POST'])
def search():
    brand_name = request.form.get('brand_name').lower()
    curr_month = datetime.now().strftime("%B_%Y")
    # Load the Excel file and filter by brand name
    df=pd.read_excel(f'artifacts/data/stock_data/{curr_month}.xlsx')
    filtered_data = df[df['Brand'].str.contains(brand_name, case=False, na=False)]

    # Convert the filtered data to a list of dictionaries
    products = filtered_data.to_dict(orient='records')

    return render_template('admin_dashboard.html', products=products, brand_name=brand_name)
           

@app.route('/get_products', methods=['GET'])
def get_products():
    brand_query = request.args.get('brand', '').lower()

    if not brand_query:
        return jsonify([])  # Return an empty list if no brand is provided

    # Load data from the Excel file
    curr_month = datetime.now().strftime("%B_%Y")    
    df=pd.read_excel(f'artifacts/data/stock_data/{curr_month}.xlsx')

    # Adjust the file path if needed

    # Filter products based on the brand
    filtered_df = df[df['Brand'].str.lower() == brand_query]

    # Get product names
    products = filtered_df['Product'].tolist()
    return jsonify(products)

AWS_ACCESS_KEY = os.environ.get("AWS_ACCESS_KEY_ID")
AWS_SECRET_KEY = os.environ.get("AWS_SECRET_ACCESS_KEY")
AWS_BUCKET_NAME = os.environ.get("S3_BUCKET_NAME")
S3_REGION = os.environ.get("AWS_REGION")

# Initialize S3 Client
s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY,
    region_name=S3_REGION,
)

def upload_to_s3( file_path,bucket_name,s3_file_path):
    try:
        s3_client.upload_file(file_path,bucket_name,s3_file_path)
        return True
    except Exception as e:
        print(f"Error uploading to S3: {e}")
        return False
    
def create_transaction_log():
    curr_month=datetime.now().strftime("%B_%Y")
    file_path=os.path.join(f'./artifacts/data/transaction_log/{curr_month}.xlsx')

    columns=['Code', 'Product', 'Sold', 'Purchased', 'Consumer', 'Date', 'Remarks']
    brand_names=['hind','zume','earthco','glass','2d','kuber','bio','gcdc']
    df=pd.DataFrame(columns=columns)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=brand_names[0], index=False)
        df.to_excel(writer, sheet_name=brand_names[1], index=False)
        df.to_excel(writer, sheet_name=brand_names[2], index=False)
        df.to_excel(writer, sheet_name=brand_names[3], index=False)
        df.to_excel(writer, sheet_name=brand_names[4], index=False)
        df.to_excel(writer, sheet_name=brand_names[5], index=False)
        df.to_excel(writer, sheet_name=brand_names[6], index=False)
        df.to_excel(writer, sheet_name=brand_names[7], index=False)

@app.route("/upload-files", methods=["POST"])
def upload_files():
    try:
        first_day_of_curr_month = datetime.now().replace(day=1)

        last_day_of_prev_month = first_day_of_curr_month - timedelta(days=1)

        prev_month_name = last_day_of_prev_month.strftime("%B_%Y")

        stock_data_path = os.path.join( f"artifacts/data/stock_data/{prev_month_name}.xlsx")
        transaction_log_path = os.path.join(f"artifacts/data/transaction_log/{prev_month_name}.xlsx")

        upload_to_s3(stock_data_path,AWS_BUCKET_NAME,f'stock_data/{prev_month_name}.xlsx')
        upload_to_s3(transaction_log_path,AWS_BUCKET_NAME,f'transaction_data/{prev_month_name}.xlsx')
        
        stock_data = pd.read_excel(stock_data_path)

        # Update the stock data file: set opening stock = closing stock
        stock_data["Opening Stock"] = stock_data["Closing Stock"]
        stock_data["Sale Out"]=0
        stock_data['Purchase']=0

        # Save modified files temporarily
        curr_month = datetime.now().strftime("%B_%Y")

        # Save stock data file
        stock_data.to_excel(os.path.join( f"artifacts/data/stock_data/{curr_month}.xlsx"), index=False)
        create_transaction_log()


        return jsonify({"message": "Files uploaded successfully!"}), 200
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500




if __name__=="__main__":
    app.run(host="0.0.0.0",port=8080,debug=True)  
